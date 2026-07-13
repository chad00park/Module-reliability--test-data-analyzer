#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module Reliability Data Analyzer
================================
Discrete Reliability Data Analyzer 기반 + Phase 차원 확장
- CONT_XX 표식으로 phase 자동 인식 (phase 개수/item 구성은 신뢰성마다 가변)
- Phase 밖 item(단위 있는 것)은 phase 무관 항목으로 별도 분석
- Read-out/Delta % graph: item당 1개, X축은 phase별 시료 구간을 이어 배치
- Box plot: 'Phase별 분석' / 'Test item별 분석' 두 모드 (PDF에는 둘 다 포함)
"""

import csv
import math
import os
import re
import sys
import threading
import traceback

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser, simpledialog

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.transforms as mtransforms
from matplotlib import font_manager as _fm
for _f in ("Malgun Gothic", "AppleGothic", "NanumGothic"):
    if any(_f == f.name for f in _fm.fontManager.ttflist):
        matplotlib.rcParams["font.family"] = _f
        matplotlib.rcParams["axes.unicode_minus"] = False
        break

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except Exception:
    HAS_DND = False

MAX_SAMPLES = 500
MAX_PARAMS = 500
MAX_READOUTS = 20

READOUT_COLORS = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
    "#aec7e8", "#ffbb78", "#98df8a", "#ff9896", "#c5b0d5",
    "#c49c94", "#f7b6d2", "#c7c7c7", "#dbdb8d", "#9edae5",
]

NO_PHASE = "(공통)"   # phase 밖 item의 내부 표시명


def col_title(col):
    """컬럼명 'ITEM@BIAS (UNIT)[ #n]' → 'ITEM [UNIT] @BIAS[ #n]' 표시 형식."""
    m = re.match(r"^(.*?)@(.*?)\s*\(([^()]*)\)(\s*#\d+)?$", col)
    if not m:
        return col
    item, bias, unit, suf = m.group(1), m.group(2), m.group(3), m.group(4) or ""
    return f"{item} [{unit}] @{bias}{suf}"


def center_window(win, w=None, h=None):
    win.update_idletasks()
    ww = w or win.winfo_width() or win.winfo_reqwidth()
    wh = h or win.winfo_height() or win.winfo_reqheight()
    x = (win.winfo_screenwidth() - ww) // 2
    y = (win.winfo_screenheight() - wh) // 2
    win.geometry(f"+{x}+{y}")


# ============================================================================
# 1. 파일명 파싱 (discrete와 동일)
# ============================================================================
READOUT_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*(hr|hrs|h|hour|hours|cyc|cycle|cycles|cy)$", re.I)
LOT_RE = re.compile(r"^lot\s*([A-Za-z0-9]+)$", re.I)


def parse_filename(path):
    base = os.path.splitext(os.path.basename(path))[0]
    tokens = [t for t in re.split(r"[_\s]+", base) if t]
    readout_label = None
    readout_value = None
    lot = None
    rest = []
    for tok in tokens:
        m = READOUT_RE.match(tok)
        if m and readout_label is None:
            readout_value = float(m.group(1))
            unit = m.group(2).lower()
            unit = "hr" if unit.startswith("h") else "cyc"
            readout_label = f"{m.group(1)}{unit}"
            continue
        m = LOT_RE.match(tok)
        if m and lot is None:
            lot = "LOT" + m.group(1).upper()
            continue
        rest.append(tok)
    if readout_label is None or lot is None or not rest:
        raise ValueError(
            f"파일명 인식 실패: '{os.path.basename(path)}'\n"
            "파일 이름은 신뢰성명 + Lot번호 + Read-out 형식이어야 합니다.\n"
            "예: HTRB_Lot1_0hr, HTBG+_Lot2_500cyc (구분자는 _ 또는 공백)"
        )
    return "_".join(rest).upper(), lot, readout_label, readout_value


# ============================================================================
# 2. 데이터 파싱 — discrete 좌표 규칙 + Phase 분할
# ============================================================================
def _read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".csv":
        for enc in ("utf-8-sig", "cp949", "latin-1"):
            try:
                with open(path, newline="", encoding=enc) as f:
                    rows = [[(c or "").strip() for c in r] for r in csv.reader(f)]
                break
            except UnicodeDecodeError:
                rows = []
                continue
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c)).strip() for c in r])
        wb.close()
    else:
        raise ValueError(f"지원하지 않는 확장자: {ext}")
    return rows


def _cell(row, idx):
    return row[idx].strip() if idx < len(row) else ""


PHASE_RE = re.compile(r"^CONT_(\w+)$", re.I)


def parse_data_file(path):
    """Module 파일 파싱.
    반환: (phases, columns, data)
      phases : 파일 등장 순서의 phase 이름 목록 (예: ['BL','BH',...])
      columns: [(phase, colname)] — phase는 NO_PHASE 또는 phase명.
               colname은 'ITEM@BIAS (UNIT)[ #n]' (phase 내 중복 시 접미사)
      data   : {sample_no: {(phase, colname): float|None}}"""
    fname = os.path.basename(path)
    rows = _read_rows(path)

    item_row = bias_row = unit_row = None
    data_start = None
    for i, row in enumerate(rows):
        c6 = _cell(row, 6)
        if item_row is None and c6 == "Item":
            item_row = row  # 첫 번째 Item 행만 사용
        elif c6 == "Bias1" and bias_row is None:
            bias_row = row
        elif c6 == "Unit" and unit_row is None:
            unit_row = row
        if data_start is None and _cell(row, 0) == "Test No.":
            data_start = i + 1

    missing = [n for n, v in
               [("Item", item_row), ("Bias1", bias_row),
                ("Unit", unit_row), ("Test No.", data_start)] if v is None]
    if missing:
        raise ValueError(f"'{fname}' 에서 {', '.join(missing)} 행을 찾을 수 없습니다.")

    ncol = max(len(item_row), len(bias_row), len(unit_row))
    phases = []
    columns = []            # [(phase, colname)] 등장 순서
    col_map = []            # [((phase, colname), file_col_index)]
    cur_phase = NO_PHASE
    name_count = {}         # (phase, base_name) -> count
    for j in range(7, ncol):
        item = _cell(item_row, j)
        if not item:
            continue
        pm = PHASE_RE.match(item)
        if pm:
            cur_phase = pm.group(1).upper()
            if cur_phase not in phases:
                phases.append(cur_phase)
            continue  # CONT_XX 자체는 데이터 대상 아님
        unit = _cell(unit_row, j)
        if not unit:
            continue  # 단위 빈 item은 모든 그래프에서 제외
        bias = _cell(bias_row, j)
        base = f"{item}@{bias} ({unit})"
        key = (cur_phase, base)
        if key in name_count:
            name_count[key] += 1
            colname = f"{base} #{name_count[key]}"
        else:
            name_count[key] = 1
            colname = base
        columns.append((cur_phase, colname))
        col_map.append(((cur_phase, colname), j))

    data = {}
    for row in rows[data_start:]:
        s = _cell(row, 0)
        if not s:
            continue
        try:
            sample = int(float(s))
        except ValueError:
            continue
        vals = {}
        for key, j in col_map:
            v = _cell(row, j)
            try:
                vals[key] = float(v)
            except ValueError:
                vals[key] = None
        data[sample] = vals

    return phases, columns, data


# ============================================================================
# 3. 데이터 모델 — Lot 그룹 × (phase, col) 키
# ============================================================================
class DataModel:
    """
    g[lot] = {
        'readouts': 정렬된 Read-out 목록,
        'phases'  : phase 목록 (파일 등장 순서),
        'rep_cols': phase 내 반복 item 컬럼명 목록 (등장 순서, phase 간 합집합),
        'np_cols' : phase 밖(공통) 컬럼명 목록,
        'samples' : 정렬된 시료 번호,
        'data'    : data[readout][sample][(phase, colname)] = float|None
    }
    편집 키: (lot, readout, colname, phase, sample)
    """

    def __init__(self):
        self.reliability = None
        self.groups = []
        self.g = {}
        self.deleted = set()
        self.color_over = {}
        self.ylim = {}          # (lot, colname) -> (min,max)
        self.ref_lines = {}     # (lot, colname, kind) -> [ {axis,value,name,color} ]
        self._undo = []
        self._redo = []

    # ---- 로드 ------------------------------------------------------------
    def load(self, files):
        errors = []
        raw = {}
        rels = {}
        for p in files:
            fname = os.path.basename(p)
            try:
                rel, lot, rl, rv = parse_filename(p)
                phases, cols, d = parse_data_file(p)
            except ValueError as e:
                errors.append(str(e))
                continue
            rels.setdefault(rel, []).append(fname)
            raw.setdefault(lot, []).append((rl, rv, phases, cols, d, fname))

        if not raw:
            return errors
        if len(rels) > 1:
            detail = "\n".join(f"  {r}: {', '.join(fs)}" for r, fs in rels.items())
            errors.append("한 번에 하나의 신뢰성만 분석할 수 있습니다.\n"
                          f"여러 신뢰성이 섞여 있습니다:\n{detail}")
            return errors
        self.reliability = next(iter(rels))

        self.groups = []
        self.g = {}
        for lot in sorted(raw):
            entries = raw[lot]
            seen_ro = {}
            dup = []
            for rl, rv, phases, cols, d, fname in entries:
                if rl in seen_ro:
                    dup.append(f"{lot}의 {rl}: '{seen_ro[rl]}' 와 '{fname}'")
                else:
                    seen_ro[rl] = fname
            if dup:
                errors.append("동일 Lot에 같은 Read-out 파일이 중복되었습니다:\n"
                              + "\n".join(dup))
                continue
            if len(entries) > MAX_READOUTS:
                errors.append(f"{lot}: Read-out 파일이 {MAX_READOUTS}개를 "
                              f"초과합니다 ({len(entries)}개).")
                continue

            entries.sort(key=lambda x: x[1])
            readouts = [e[0] for e in entries]
            phase_order, rep_cols, np_cols = [], [], []
            samp_union = set()
            data = {}
            for rl, rv, phases, cols, d, fname in entries:
                for ph in phases:
                    if ph not in phase_order:
                        phase_order.append(ph)
                for ph, c in cols:
                    if ph == NO_PHASE:
                        if c not in np_cols:
                            np_cols.append(c)
                    else:
                        if c not in rep_cols:
                            rep_cols.append(c)
                samp_union.update(d.keys())
                data[rl] = d

            n_params = len(rep_cols) + len(np_cols)
            if n_params > MAX_PARAMS:
                errors.append(f"{lot}: Parameter가 {MAX_PARAMS}개를 초과합니다 "
                              f"({n_params}개).")
                continue
            if len(samp_union) > MAX_SAMPLES:
                errors.append(f"{lot}: Sample이 {MAX_SAMPLES}개를 초과합니다 "
                              f"({len(samp_union)}개).")
                continue

            self.groups.append(lot)
            self.g[lot] = dict(readouts=readouts, phases=phase_order,
                               rep_cols=rep_cols, np_cols=np_cols,
                               samples=sorted(samp_union), data=data)

        self.deleted.clear()
        self.color_over.clear()
        self.ylim.clear()
        self.ref_lines.clear()
        self._undo.clear()
        self._redo.clear()
        return errors

    # ---- 접근자 ----------------------------------------------------------
    def readouts(self, lot):
        return self.g[lot]["readouts"]

    def phases(self, lot):
        return self.g[lot]["phases"]

    def rep_cols(self, lot):
        return self.g[lot]["rep_cols"]

    def np_cols(self, lot):
        return self.g[lot]["np_cols"]

    def samples(self, lot):
        return self.g[lot]["samples"]

    def col_phases(self, lot, col):
        """해당 컬럼이 그려질 phase 목록. 공통 컬럼이면 [NO_PHASE]."""
        if col in self.g[lot]["np_cols"]:
            return [NO_PHASE]
        return self.g[lot]["phases"]

    # ---- 값 조회 -----------------------------------------------------------
    def value(self, lot, readout, col, phase, sample):
        if (lot, readout, col, phase, sample) in self.deleted:
            return None
        return self.g[lot]["data"].get(readout, {}).get(sample, {}).get((phase, col))

    def seg_series(self, lot, readout, col):
        """phase 구간을 이어 붙인 시리즈.
        반환: (positions, values, seg_info)
          seg_info: [(phase, start_pos, end_pos)] — 구간 경계/이름 표기에 사용"""
        samples = self.samples(lot)
        N = len(samples)
        phs = self.col_phases(lot, col)
        xs, ys = [], []
        segs = []
        for pi, ph in enumerate(phs):
            base = pi * N
            segs.append((ph, base + 1, base + N))
            for si, s in enumerate(samples):
                xs.append(base + si + 1)
                v = self.value(lot, readout, col, ph, s)
                ys.append(math.nan if v is None else v)
        return xs, ys, segs

    def seg_delta(self, lot, readout, col):
        """초기 Read-out 대비 변화율(%) — 같은 phase 같은 시료끼리 비교."""
        base_ro = self.readouts(lot)[0]
        samples = self.samples(lot)
        N = len(samples)
        phs = self.col_phases(lot, col)
        xs, ys = [], []
        segs = []
        for pi, ph in enumerate(phs):
            base = pi * N
            segs.append((ph, base + 1, base + N))
            for si, s in enumerate(samples):
                xs.append(base + si + 1)
                v0 = self.value(lot, base_ro, col, ph, s)
                v = self.value(lot, readout, col, ph, s)
                if v0 is None or v is None or v0 == 0:
                    ys.append(math.nan)
                else:
                    ys.append((v - v0) / v0 * 100.0)
        return xs, ys, segs

    def pos_to_phase_sample(self, lot, col, pos_index):
        """seg_series의 인덱스 → (phase, sample)."""
        samples = self.samples(lot)
        N = len(samples)
        phs = self.col_phases(lot, col)
        pi = pos_index // N
        si = pos_index % N
        if pi >= len(phs):
            return None, None
        return phs[pi], samples[si]

    def box_values(self, lot, readout, col, phase=None):
        """phase 지정 시 해당 phase만, None이면 모든 phase stack."""
        phs = [phase] if phase is not None else self.col_phases(lot, col)
        out = []
        for ph in phs:
            for s in self.samples(lot):
                v = self.value(lot, readout, col, ph, s)
                if v is not None:
                    out.append(v)
        return out

    # ---- diff 기반 Undo/Redo ------------------------------------------------
    def _apply(self, action, forward=True):
        kind = action[0]
        if kind == "del":
            _, keys = action
            if forward:
                self.deleted.update(keys)
            else:
                self.deleted.difference_update(keys)
        elif kind == "color":
            _, key, old, new = action
            c = new if forward else old
            if c is None:
                self.color_over.pop(key, None)
            else:
                self.color_over[key] = c
        elif kind == "ylim":
            _, key, old, new = action
            v = new if forward else old
            if v is None:
                self.ylim.pop(key, None)
            else:
                self.ylim[key] = v
        elif kind == "refline":
            _, key, old, new = action
            v = new if forward else old
            if not v:
                self.ref_lines.pop(key, None)
            else:
                self.ref_lines[key] = v

    def do(self, action):
        self._apply(action, True)
        self._undo.append(action)
        self._redo.clear()

    def undo(self):
        if not self._undo:
            return False
        a = self._undo.pop()
        self._apply(a, False)
        self._redo.append(a)
        return True

    def redo(self):
        if not self._redo:
            return False
        a = self._redo.pop()
        self._apply(a, True)
        self._undo.append(a)
        return True

    # ---- 편집 액션 -----------------------------------------------------------
    def delete_point(self, lot, readout, col, phase, sample):
        self.do(("del", frozenset({(lot, readout, col, phase, sample)})))

    def delete_sample_all_readouts(self, lot, col, phase, sample):
        keys = frozenset((lot, r, col, phase, sample) for r in self.readouts(lot))
        self.do(("del", keys))

    def set_color(self, lot, readout, col, phase, sample, color):
        key = (lot, readout, col, phase, sample)
        self.do(("color", key, self.color_over.get(key), color))

    def set_ref_lines(self, lot, col, kind, lines):
        key = (lot, col, kind)
        old = list(self.ref_lines.get(key, []))
        self.do(("refline", key, old, list(lines)))

    def set_ylim(self, lot, col, ymin, ymax):
        key = (lot, col)
        self.do(("ylim", key, self.ylim.get(key),
                 None if ymin is None else (ymin, ymax)))

    # ---- 통계 -----------------------------------------------------------------
    def stats(self, lot, readout, col, phase=None):
        vals = self.box_values(lot, readout, col, phase)
        n = len(vals)
        if n == 0:
            return dict(SS=0, Min=math.nan, Max=math.nan, AVG=math.nan, STD=math.nan)
        avg = sum(vals) / n
        std = (sum((v - avg) ** 2 for v in vals) / (n - 1)) ** 0.5 if n > 1 else 0.0
        return dict(SS=n, Min=min(vals), Max=max(vals), AVG=avg, STD=std)


# ============================================================================
# 4. 그래프 렌더링
# ============================================================================
def draw_ref_lines(ax, model, group, col, kind):
    """(점선) 임의 기준선 + 이름을 그래프 안쪽에 표시.
    이름은 data line과 겹침이 적은 쪽(위/아래, 좌/우)에 배치하고
    기준선과 가는 실선으로 연결한다."""
    lines = model.ref_lines.get((group, col, kind), [])
    if not lines:
        return
    # 현재 그려진 data line들의 점 수집 (겹침 회피 판단용)
    pts_x, pts_y = [], []
    for ln in ax.get_lines():
        if ln.get_marker() == "o":
            xd = ln.get_xdata()
            yd = ln.get_ydata()
            for x, y in zip(xd, yd):
                try:
                    if not (math.isnan(float(y))):
                        pts_x.append(float(x))
                        pts_y.append(float(y))
                except (TypeError, ValueError):
                    pass
    y0, y1 = ax.get_ylim()
    x0, x1 = ax.get_xlim()
    yr = (y1 - y0) or 1.0
    xr = (x1 - x0) or 1.0
    for rl in lines:
        c = rl.get("color") or "red"
        name = rl.get("name", "")
        v = rl["value"]
        if rl["axis"] == "y":
            ax.axhline(v, color=c, ls=":", lw=1.2, zorder=4)
            if not name:
                continue
            dy = 0.07 * yr
            # 선 위/아래 근처의 data 점 수를 비교해 덜 붐비는 쪽 선택
            above = sum(1 for y in pts_y if v < y <= v + 2 * dy)
            below = sum(1 for y in pts_y if v - 2 * dy <= y < v)
            ty = v + dy if above <= below else v - dy
            ty = min(max(ty, y0 + 0.03 * yr), y1 - 0.03 * yr)
            tx = x1 - 0.02 * xr          # 이름 위치 (그래프 안쪽 우측)
            ax_anchor = x1 - 0.10 * xr   # 기준선 위 연결점
            ax.annotate(name, xy=(ax_anchor, v), xytext=(tx, ty),
                        ha="right", va="center", fontsize=7, color=c,
                        zorder=6, arrowprops=dict(arrowstyle="-", color=c,
                                                  lw=0.7, alpha=0.9))
        else:
            ax.axvline(v, color=c, ls=":", lw=1.2, zorder=4)
            if not name:
                continue
            dx = 0.05 * xr
            right = sum(1 for x, y in zip(pts_x, pts_y)
                        if v < x <= v + 2 * dx and y > y1 - 0.25 * yr)
            left = sum(1 for x, y in zip(pts_x, pts_y)
                       if v - 2 * dx <= x < v and y > y1 - 0.25 * yr)
            tx = v + dx if right <= left else v - dx
            tx = min(max(tx, x0 + 0.03 * xr), x1 - 0.03 * xr)
            ty = y1 - 0.08 * yr          # 이름 위치 (그래프 안쪽 상단)
            ay_anchor = y1 - 0.16 * yr   # 기준선 위 연결점
            ax.annotate(name, xy=(v, ay_anchor), xytext=(tx, ty),
                        ha="center", va="bottom", fontsize=7, color=c,
                        zorder=6, arrowprops=dict(arrowstyle="-", color=c,
                                                  lw=0.7, alpha=0.9))


def readout_color(model, lot, readout):
    return READOUT_COLORS[model.readouts(lot).index(readout) % len(READOUT_COLORS)]


def graph_title(model, lot, col, phase=None):
    t = f"{model.reliability}: {lot}, {col_title(col)}"
    if phase is not None and phase != NO_PHASE:
        t += f" ({phase})"
    return t


def delta_title(model, lot, col):
    t = f"{model.reliability}: {lot}, {col_title(col)}"
    return re.sub(r"\[[^\[\]]*\]", "[%]", t, count=1)


def _seg_axis(ax, model, lot, col, segs, screen=False):
    """구간형 X축: 시료 라벨(짝수) 반복 + phase 경계 점선 + phase 이름 표기.
    screen=True(프로그램 화면): tick 폰트 +2, phase 이름을 축에 근접 배치(고정 픽셀).
    screen=False(PDF): 기존 배치 유지."""
    samples = model.samples(lot)
    N = len(samples)
    multi = len(segs) > 1 or segs[0][0] != NO_PHASE
    ticks, labels = [], []
    minor_ticks = []
    for pi, (ph, start, end) in enumerate(segs):
        base = start - 1
        for si, s in enumerate(samples):
            if s % 2 == 0:
                ticks.append(base + si + 1)
                labels.append(str(s))
            else:
                minor_ticks.append(base + si + 1)  # 홀수 시료는 minor tick
        if pi > 0:
            ax.axvline(start - 0.5, color="gray", lw=0.9, ls="--", alpha=0.8)
    ax.set_xticks(ticks)
    ax.set_xticklabels(labels)
    ax.set_xticks(minor_ticks, minor=True)
    ax.set_xlim(0, segs[-1][2] + 1)
    ax.tick_params(labelsize=8 if screen else 6)
    if multi:
        for ph, start, end in segs:
            name = "" if ph == NO_PHASE else ph
            if screen:
                # 화면: tick 라벨 바로 아래 고정 픽셀 오프셋 (축 크기 무관하게 근접)
                ax.annotate(name, xy=((start + end) / 2, 0),
                            xycoords=("data", "axes fraction"),
                            xytext=(0, -18), textcoords="offset points",
                            ha="center", va="top", fontsize=9, fontweight="bold",
                            annotation_clip=False)
            else:
                trans = mtransforms.blended_transform_factory(ax.transData, ax.transAxes)
                ax.text((start + end) / 2, -0.30, name, transform=trans,
                        ha="center", va="top", fontsize=8, fontweight="bold")
        # phase 이름이 축 설명을 겸하므로 Sample No. 라벨은 생략 (겹침 방지)
    else:
        ax.set_xlabel("Sample No.", fontsize=8)


def draw_line(ax, model, lot, col, picker=False, screen=False):
    artists = {}
    segs = None
    for r in model.readouts(lot):
        c = readout_color(model, lot, r)
        xs, ys, segs = model.seg_series(lot, r, col)
        kw = dict(marker="o", ms=3.5, lw=0.9, color=c, label=r)
        if picker:
            kw["picker"] = 5
        line, = ax.plot(xs, ys, **kw)
        artists[r] = line
        # 색 변경 마커(삼각형)
        for idx, (x, v) in enumerate(zip(xs, ys)):
            ph, s = model.pos_to_phase_sample(lot, col, idx)
            oc = model.color_over.get((lot, r, col, ph, s))
            if oc and not math.isnan(v):
                ax.plot([x], [v], marker="^", ms=5, color=oc, ls="none", zorder=5)
    ax.set_title(graph_title(model, lot, col), fontsize=9)
    mu = re.search(r"\(([^()]*)\)\s*(?:#\d+)?$", col)
    unit = f" ({mu.group(1)})" if mu else ""
    ax.set_ylabel(col.split("@")[0] + unit, fontsize=8)
    _seg_axis(ax, model, lot, col, segs, screen=screen)
    if (lot, col) in model.ylim:
        ax.set_ylim(*model.ylim[(lot, col)])
    else:
        # 기본: 모든 Read-out 그래프는 Y축 0 ~ (데이터 max의 110%)
        vmax = None
        for _ln in ax.get_lines():
            for _y in _ln.get_ydata():
                try:
                    fy = float(_y)
                except (TypeError, ValueError):
                    continue
                if not math.isnan(fy) and (vmax is None or fy > vmax):
                    vmax = fy
        if vmax is not None and vmax > 0:
            ax.set_ylim(0, vmax * 1.10)
        else:
            ax.set_ylim(bottom=0)
    draw_ref_lines(ax, model, lot, col, "line")
    ax.legend(fontsize=9 if screen else 6, ncol=2)
    ax.grid(True, alpha=0.3)
    return artists


def draw_delta(ax, model, lot, col, screen=False):
    segs = None
    for r in model.readouts(lot)[1:]:
        c = readout_color(model, lot, r)
        xs, ys, segs = model.seg_delta(lot, r, col)
        ax.plot(xs, ys, marker="o", ms=3.5, lw=0.9, color=c, label=r)
        # 색 변경된 점은 Delta 그래프에도 삼각형으로 동일 표시
        for idx, (x, v) in enumerate(zip(xs, ys)):
            ph, s = model.pos_to_phase_sample(lot, col, idx)
            oc = model.color_over.get((lot, r, col, ph, s))
            if oc and not math.isnan(v):
                ax.plot([x], [v], marker="^", ms=5, color=oc, ls="none", zorder=5)
    if segs is None:
        xs, ys, segs = model.seg_delta(lot, model.readouts(lot)[0], col)
    ax.set_title(delta_title(model, lot, col), fontsize=9)
    ax.set_ylabel("Delta (%)", fontsize=8)
    _seg_axis(ax, model, lot, col, segs, screen=screen)
    ax.axhline(0, color="gray", lw=0.8, ls="--", alpha=0.7)
    draw_ref_lines(ax, model, lot, col, "delta")
    if len(model.readouts(lot)) > 1:
        ax.legend(fontsize=9 if screen else 6, ncol=2)
    else:
        ax.text(0.5, 0.5, "Read-out 1개 — Delta 없음", transform=ax.transAxes,
                ha="center", va="center", fontsize=9, color="gray")
    ax.grid(True, alpha=0.3)


def draw_box(ax, model, lot, col, phase=None, stacked=False,
             picker=False, stats_table=True):
    """Box plot.
    phase 지정: 해당 phase만 / stacked=True: 모든 phase stack /
    공통 컬럼: phase=NO_PHASE로 호출."""
    readouts = model.readouts(lot)
    ph_arg = None if stacked else phase
    groups_vals = [model.box_values(lot, r, col, ph_arg) for r in readouts]
    try:
        bp = ax.boxplot(groups_vals, tick_labels=readouts, patch_artist=True,
                        showfliers=True, medianprops=dict(color="black"))
    except TypeError:
        bp = ax.boxplot(groups_vals, labels=readouts, patch_artist=True,
                        showfliers=True, medianprops=dict(color="black"))
    for patch, r in zip(bp["boxes"], readouts):
        patch.set_facecolor(readout_color(model, lot, r))
        patch.set_alpha(0.5)
    for fl, r in zip(bp["fliers"], readouts):
        fl.set(marker="o", markerfacecolor=readout_color(model, lot, r),
               markeredgecolor="black", markersize=3)
        if picker:
            fl.set_picker(5)
    # 색 변경 삼각형 — 모든 Box 모드에서 표시 (stacked 포함)
    for (lt, r, c, ph, s), oc in model.color_over.items():
        if lt != lot or c != col:
            continue
        if not stacked and phase is not None and ph != phase:
            continue
        v = model.value(lot, r, c, ph, s)
        if v is not None and r in readouts:
            ax.plot([readouts.index(r) + 1], [v], marker="^", ms=5,
                    color=oc, ls="none", zorder=5)
    title_phase = None if stacked else phase
    ax.set_title(graph_title(model, lot, col, title_phase), fontsize=8)
    ax.tick_params(labelsize=7)
    if (lot, col) in model.ylim:
        ax.set_ylim(*model.ylim[(lot, col)])
    draw_ref_lines(ax, model, lot, col, "box")
    ax.grid(True, alpha=0.3)

    if stats_table:
        trans = mtransforms.blended_transform_factory(ax.transData, ax.transAxes)
        ax.tick_params(axis="x", labelbottom=False)
        y0, dy, fs = -0.07, 0.075, 8
        for i, r in enumerate(readouts):
            ax.text(i + 1, y0, r, transform=trans,
                    ha="center", va="center", fontsize=fs, fontweight="bold")
        rows = [("S/S", "SS"), ("Min", "Min"), ("Max", "Max"),
                ("AVG", "AVG"), ("STD", "STD")]
        for k, (label, key) in enumerate(rows):
            y = y0 - dy * (k + 1)
            ax.text(-0.01, y, label, transform=ax.transAxes,
                    ha="right", va="center", fontsize=fs, fontweight="bold")
            for i, r in enumerate(readouts):
                st = model.stats(lot, r, col, ph_arg)
                v = st[key]
                txt = str(v) if key == "SS" else f"{v:.4g}"
                ax.text(i + 1, y, txt, transform=trans,
                        ha="center", va="center", fontsize=fs)
    else:
        ax.set_xlabel("Read-out", fontsize=8)
    return bp


# ============================================================================
# 5. PDF 출력 — Lot별: [item 쌍 4줄/페이지] → [Box Phase별 3x2] → [Box item별 3x2]
# ============================================================================
PPT_LANDSCAPE = (13.33, 7.5)


def export_pdf(model, pairs, path, include_phase=False, progress_cb=None):
    """pairs: [(lot, col)]. col은 반복 item 또는 공통 item."""
    per_lot = {}
    for g, c in pairs:
        per_lot.setdefault(g, []).append(c)
    order = [g for g in model.groups if g in per_lot]

    # 총 페이지 수 계산
    total = 0
    plan = {}
    for g in order:
        cols = per_lot[g]
        phase_units = []
        if include_phase:
            for c in cols:
                for ph in model.col_phases(g, c):
                    phase_units.append((c, ph))
        n_pair = math.ceil(len(cols) / 2)
        n_boxA = math.ceil(len(phase_units) / 6) if include_phase else 0
        n_boxB = math.ceil(len(cols) / 6)
        plan[g] = (cols, phase_units, n_pair, n_boxA, n_boxB)
        total += n_pair + n_boxA + n_boxB

    done = 0
    with PdfPages(path) as pdf:
        for g in order:
            cols, phase_units, n_pair, n_boxA, n_boxB = plan[g]
            head = f"{model.reliability}: {g}"
            # 1) item별 Read-out + Delta % 쌍 (1페이지 1줄1개 x 4줄 = item 2개분)
            for i in range(0, len(cols), 2):
                fig = Figure(figsize=PPT_LANDSCAPE)
                fig.suptitle(f"{head} — Read-out & Delta %", fontsize=12)
                for k in range(2):
                    ax1 = fig.add_subplot(4, 1, 2 * k + 1)
                    ax2 = fig.add_subplot(4, 1, 2 * k + 2)
                    if i + k < len(cols):
                        c = cols[i + k]
                        draw_line(ax1, model, g, c)
                        draw_delta(ax2, model, g, c)
                    else:
                        ax1.axis("off")
                        ax2.axis("off")
                fig.tight_layout(rect=(0, 0, 1, 0.96))
                pdf.savefig(fig)
                done += 1
                if progress_cb:
                    progress_cb(done, total)
            # 2) Box plot — Test item별 분석 (기본, phase stack, 3개 x 2줄)
            for i in range(0, len(cols), 6):
                fig = Figure(figsize=PPT_LANDSCAPE)
                fig.suptitle(f"{head} — Box plot (by Test item)", fontsize=12)
                for k in range(6):
                    ax = fig.add_subplot(2, 3, k + 1)
                    if i + k < len(cols):
                        c = cols[i + k]
                        if c in model.np_cols(g):
                            draw_box(ax, model, g, c, phase=NO_PHASE, stats_table=True)
                        else:
                            draw_box(ax, model, g, c, stacked=True, stats_table=True)
                    else:
                        ax.axis("off")
                fig.subplots_adjust(top=0.90, bottom=0.15, left=0.07, right=0.98,
                                    hspace=0.85, wspace=0.35)
                pdf.savefig(fig)
                done += 1
                if progress_cb:
                    progress_cb(done, total)
            # 3) Box plot — Phase별 분석 (선택 시에만 포함, 3개 x 2줄)
            for i in range(0, len(phase_units), 6):
                fig = Figure(figsize=PPT_LANDSCAPE)
                fig.suptitle(f"{head} — Box plot (by Phase)", fontsize=12)
                for k in range(6):
                    ax = fig.add_subplot(2, 3, k + 1)
                    if i + k < len(phase_units):
                        c, ph = phase_units[i + k]
                        draw_box(ax, model, g, c, phase=ph, stats_table=True)
                    else:
                        ax.axis("off")
                fig.subplots_adjust(top=0.90, bottom=0.15, left=0.07, right=0.98,
                                    hspace=0.85, wspace=0.35)
                pdf.savefig(fig)
                done += 1
                if progress_cb:
                    progress_cb(done, total)


# ============================================================================
# 6. GUI
# ============================================================================
BaseTk = TkinterDnD.Tk if HAS_DND else tk.Tk


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("Module Reliability Data Analyzer")
        self.geometry("1250x850")
        center_window(self, 1250, 850)
        self.model = DataModel()
        self.files = []
        self.selected = []   # [(lot, col)]
        self.cur_idx = 0
        self.box_mode = tk.StringVar(value="item")    # 'item'(기본) | 'phase'
        self._phase_mode_confirmed = False
        self.box_page = 0
        self._build_start()

    def _reset_all(self):
        """진행 내용을 모두 초기화하고 처음 화면으로."""
        if not messagebox.askyesno("처음으로", "진행하던 내용을 모두 지우고\n처음 화면으로 돌아갈까요?",
                                   parent=self):
            return
        self.model = DataModel()
        self.files = []
        self.selected = []
        self.cur_idx = 0
        self.box_page = 0
        self.box_mode.set("item")
        self._phase_mode_confirmed = False
        self._build_start()

    # ---- 화면 1: 파일 선택 -------------------------------------------------
    def _build_start(self):
        for w in self.winfo_children():
            w.destroy()
        frm = ttk.Frame(self, padding=20)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Module Reliability Data Analyzer",
                  font=("", 16, "bold")).pack(pady=10)
        ttk.Label(frm, text="파일 이름은 신뢰성명 + Lot번호 + Read-out 형식이어야 합니다.\n"
                            "예: HTRB_Lot1_0hr.csv, HTBG+_Lot2_500cyc.xlsx (구분자: _ 또는 공백)\n"
                            "서로 다른 Lot은 자동 구분되며, 다른 신뢰성이 섞이면 경고가 표시됩니다.").pack(pady=5)

        drop = tk.Label(frm, text="여기에 파일을 Drag && Drop 하세요"
                        if HAS_DND else "Browse 버튼으로 파일을 선택하세요",
                        relief="ridge", height=6, bg="#f0f0f0")
        drop.pack(fill="x", pady=10)
        if HAS_DND:
            drop.drop_target_register(DND_FILES)
            drop.dnd_bind("<<Drop>>", lambda e: self._add_files(self.tk.splitlist(e.data)))

        ttk.Button(frm, text="Browse...", command=self._browse).pack()
        self.file_list = tk.Listbox(frm, height=8)
        self.file_list.pack(fill="both", expand=True, pady=10)
        btns = ttk.Frame(frm)
        btns.pack()
        ttk.Button(btns, text="선택 제거", command=self._remove_file).pack(side="left", padx=5)
        ttk.Button(btns, text="다음 (파일 읽기)", command=self._load_files).pack(side="left", padx=5)

    def _browse(self):
        paths = filedialog.askopenfilenames(
            filetypes=[("Data files", "*.csv *.xlsx *.xlsm"), ("All", "*.*")])
        self._add_files(paths)

    def _add_files(self, paths):
        for p in paths:
            p = p.strip("{}")
            if p and p not in self.files:
                self.files.append(p)
                self.file_list.insert("end", os.path.basename(p))

    def _remove_file(self):
        for i in reversed(self.file_list.curselection()):
            self.file_list.delete(i)
            del self.files[i]

    def _load_files(self):
        if not self.files:
            messagebox.showwarning("알림", "파일을 먼저 선택하세요.", parent=self)
            return
        errors = self.model.load(self.files)
        if errors:
            messagebox.showerror("파일 오류", "\n\n".join(errors), parent=self)
        if not self.model.groups:
            return
        self._build_param_select()

    # ---- 화면 2: Parameter 선택 ----------------------------------------------
    def _build_param_select(self):
        for w in self.winfo_children():
            w.destroy()
        frm = ttk.Frame(self, padding=20)
        frm.pack(fill="both", expand=True)
        head = [f"신뢰성: {self.model.reliability}"]
        for g in self.model.groups:
            head.append(f"{g} (Read-out: {', '.join(self.model.readouts(g))}, "
                        f"Phase: {', '.join(self.model.phases(g)) or '없음'}, "
                        f"Sample: {len(self.model.samples(g))})")
        ttk.Label(frm, text="\n".join(head), font=("", 10, "bold"),
                  justify="left").pack(pady=5)
        ttk.Label(frm, text="분석할 Parameter를 선택하세요 (Ctrl/Shift 다중 선택)\n"
                            "※ (공통) 표기는 phase 밖 item입니다.").pack()
        self.param_lb = tk.Listbox(frm, selectmode="extended")
        self._pairs = []
        multi = len(self.model.groups) > 1
        for g in self.model.groups:
            for c in self.model.rep_cols(g):
                self._pairs.append((g, c))
                label = col_title(c)
                self.param_lb.insert("end", f"{g} | {label}" if multi else label)
            for c in self.model.np_cols(g):
                self._pairs.append((g, c))
                label = f"{col_title(c)} {NO_PHASE}"
                self.param_lb.insert("end", f"{g} | {label}" if multi else label)
        self.param_lb.pack(fill="both", expand=True, pady=10)
        btns = ttk.Frame(frm)
        btns.pack()
        ttk.Button(btns, text="전체 선택",
                   command=lambda: self.param_lb.select_set(0, "end")).pack(side="left", padx=5)
        ttk.Button(btns, text="분석 시작", command=self._analyze).pack(side="left", padx=5)
        ttk.Button(btns, text="← 파일 다시 선택", command=self._build_start).pack(side="left", padx=5)
        ttk.Button(btns, text="⟲ 처음으로 (Reset)", command=self._reset_all).pack(side="left", padx=5)
        self.pbar = ttk.Progressbar(frm, mode="determinate")
        self.pbar.pack(fill="x", pady=5)

    def _analyze(self):
        sel = self.param_lb.curselection()
        if not sel:
            messagebox.showwarning("알림", "Parameter를 선택하세요.", parent=self)
            return
        self.selected = [self._pairs[i] for i in sel]
        self.cur_idx = 0
        self.pbar["maximum"] = len(self.selected)
        self.pbar["value"] = len(self.selected)
        self.update_idletasks()
        self._build_graphs()

    # ---- 화면 3: 그래프 -------------------------------------------------------
    def _label(self, pair):
        g, c = pair
        label = col_title(c)
        if c in self.model.np_cols(g):
            label += f" {NO_PHASE}"
        return f"{g} | {label}" if len(self.model.groups) > 1 else label

    def _build_graphs(self):
        for w in self.winfo_children():
            w.destroy()
        top = ttk.Frame(self, padding=5)
        top.pack(fill="x")
        ttk.Button(top, text="← Parameter", command=self._build_param_select).pack(side="left")
        ttk.Button(top, text="◀ 이전", command=lambda: self._nav(-1)).pack(side="left", padx=3)
        self.param_var = tk.StringVar()
        self._labels = [self._label(p) for p in self.selected]
        cb = ttk.Combobox(top, textvariable=self.param_var,
                          values=self._labels, width=55, state="readonly")
        cb.pack(side="left", padx=3)
        cb.bind("<<ComboboxSelected>>",
                lambda e: self._goto(self._labels.index(self.param_var.get())))
        ttk.Button(top, text="다음 ▶", command=lambda: self._nav(1)).pack(side="left", padx=3)
        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(top, text="Y축 Min/Max", command=self._set_ylim).pack(side="left", padx=3)
        ttk.Button(top, text="기준선", command=self._ref_line_dialog).pack(side="left", padx=3)
        ttk.Button(top, text="Undo", command=self._undo).pack(side="left", padx=3)
        ttk.Button(top, text="Redo", command=self._redo).pack(side="left", padx=3)
        ttk.Button(top, text="Export PDF", command=self._export_pdf).pack(side="right", padx=3)
        ttk.Button(top, text="⟲ 처음으로", command=self._reset_all).pack(side="right", padx=3)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True)
        self.line_tab = ttk.Frame(self.nb)
        self.delta_tab = ttk.Frame(self.nb)
        self.box_tab = ttk.Frame(self.nb)
        self.nb.add(self.line_tab, text="Read-out graph")
        self.nb.add(self.delta_tab, text="Delta % graph")
        self.nb.add(self.box_tab, text="Box plot")

        self.line_fig = Figure(figsize=(10, 5))
        self.line_canvas = FigureCanvasTkAgg(self.line_fig, self.line_tab)
        self.line_canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.line_canvas, self.line_tab)
        self.line_canvas.mpl_connect("pick_event", self._on_pick_line)

        self.delta_fig = Figure(figsize=(10, 5))
        self.delta_canvas = FigureCanvasTkAgg(self.delta_fig, self.delta_tab)
        self.delta_canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.delta_canvas, self.delta_tab)

        # Box plot 탭: 모드 선택 + 페이지 이동
        ctrl = ttk.Frame(self.box_tab)
        ctrl.pack(fill="x")
        ttk.Label(ctrl, text="분석 모드:").pack(side="left", padx=(8, 3))
        ttk.Radiobutton(ctrl, text="Test item별 분석 (기본)", variable=self.box_mode,
                        value="item", command=self._box_mode_changed).pack(side="left")
        ttk.Radiobutton(ctrl, text="Phase별 분석", variable=self.box_mode,
                        value="phase", command=self._box_mode_changed).pack(side="left", padx=8)
        ttk.Button(ctrl, text="◀", width=3,
                   command=lambda: self._box_nav(-1)).pack(side="left", padx=(20, 2))
        self.box_page_lbl = ttk.Label(ctrl, text="")
        self.box_page_lbl.pack(side="left")
        ttk.Button(ctrl, text="▶", width=3,
                   command=lambda: self._box_nav(1)).pack(side="left", padx=2)

        self.box_fig = Figure(figsize=(10, 5))
        self.box_canvas = FigureCanvasTkAgg(self.box_fig, self.box_tab)
        self.box_canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.box_canvas, self.box_tab)
        self.box_canvas.mpl_connect("pick_event", self._on_pick_box)

        self._goto(0)

    def _nav(self, d):
        nxt = self.cur_idx + d
        if nxt >= len(self.selected):
            messagebox.showinfo("알림", "마지막 item입니다.", parent=self)
            return
        if nxt < 0:
            messagebox.showinfo("알림", "처음 item입니다.", parent=self)
            return
        self._goto(nxt)

    def _goto(self, idx):
        self.cur_idx = idx
        # 현재 item이 포함된 Box 페이지로 자동 동기화
        lot, col = self.selected[idx]
        if self.box_mode.get() == "item":
            same_lot = [p for p in self.selected if p[0] == lot]
            self.box_page = same_lot.index((lot, col)) // 4
        else:
            self.box_page = 0
        self.param_var.set(self._labels[idx])
        self._redraw()

    def _box_mode_changed(self):
        if self.box_mode.get() == "phase" and not self._phase_mode_confirmed:
            lot, col = self.selected[self.cur_idx]
            n_ph = len(self.model.phases(lot)) or 1
            ok = messagebox.askyesno(
                "Phase별 분석",
                f"Phase별 분석은 item마다 phase 수({n_ph}개)만큼 그래프가 생성됩니다.\n"
                "Phase별 분석을 표시할까요?", parent=self)
            if not ok:
                self.box_mode.set("item")
                return
            self._phase_mode_confirmed = True
        self.box_page = 0
        self._redraw_box()

    def _box_units(self):
        """현재 선택 item의 Box 그래프 단위 목록 반환.
        phase 모드: 현재 col의 phase들 / item 모드: 현재 col부터 연속 4개(stack)."""
        lot, col = self.selected[self.cur_idx]
        if self.box_mode.get() == "phase":
            return [(lot, col, ph, False) for ph in self.model.col_phases(lot, col)]
        # item 모드: 같은 lot의 선택 컬럼 "전체"를 단위로 (◀▶로 4개씩 페이지 이동)
        same_lot = [p for p in self.selected if p[0] == lot]
        units = []
        for g2, c2 in same_lot:
            if c2 in self.model.np_cols(g2):
                units.append((g2, c2, NO_PHASE, False))
            else:
                units.append((g2, c2, None, True))
        return units

    def _box_nav(self, d):
        units = self._box_units()
        pages = max(1, math.ceil(len(units) / 4))
        nxt = self.box_page + d
        if nxt < 0 or nxt >= pages:
            return
        self.box_page = nxt
        self._redraw_box()

    def _redraw(self):
        lot, col = self.selected[self.cur_idx]

        self.line_fig.clear()
        ax = self.line_fig.add_subplot(111)
        self._line_artists = draw_line(ax, self.model, lot, col, picker=True, screen=True)
        self.line_fig.tight_layout()
        self.line_canvas.draw()

        self.delta_fig.clear()
        axd = self.delta_fig.add_subplot(111)
        draw_delta(axd, self.model, lot, col, screen=True)
        self.delta_fig.tight_layout()
        self.delta_canvas.draw()

        self._redraw_box()

    def _redraw_box(self):
        lot, col = self.selected[self.cur_idx]
        self.box_fig.clear()
        self._box_fliers = {}
        units = self._box_units()
        pages = max(1, math.ceil(len(units) / 4))
        self.box_page = min(self.box_page, pages - 1)
        page_units = units[self.box_page * 4:(self.box_page + 1) * 4]
        self.box_page_lbl.config(text=f"{self.box_page + 1}/{pages}")
        for k, (g2, c2, ph, stacked) in enumerate(page_units):
            ax = self.box_fig.add_subplot(2, 2, k + 1)
            bp = draw_box(ax, self.model, g2, c2, phase=ph, stacked=stacked,
                          picker=(not stacked), stats_table=True)
            if not stacked and ph is not None:
                for fl, r in zip(bp["fliers"], self.model.readouts(g2)):
                    self._box_fliers[fl] = (g2, r, c2, ph)
            if c2 == col:
                title_phase = None if stacked else ph
                ax.set_title(graph_title(self.model, g2, c2, title_phase),
                             fontsize=8, fontweight="bold")
        self.box_fig.subplots_adjust(top=0.93, bottom=0.16, left=0.10,
                                     right=0.97, hspace=0.85, wspace=0.35)
        self.box_canvas.draw()

    # ---- 편집 --------------------------------------------------------------
    def _on_pick_line(self, event):
        lot, col = self.selected[self.cur_idx]
        artist = event.artist
        readout = None
        for r, ln in self._line_artists.items():
            if ln is artist:
                readout = r
                break
        if readout is None or not len(event.ind):
            return
        ph, sample = self.model.pos_to_phase_sample(lot, col, event.ind[0])
        if ph is None:
            return
        self._point_menu(lot, readout, col, ph, sample)

    def _on_pick_box(self, event):
        info = getattr(self, "_box_fliers", {}).get(event.artist)
        if info is None or not len(event.ind):
            return
        lot, readout, col, ph = info
        yval = event.artist.get_ydata()[event.ind[0]]
        sample = None
        best = float("inf")
        for s in self.model.samples(lot):
            v = self.model.value(lot, readout, col, ph, s)
            if v is None:
                continue
            d = abs(v - yval)
            if d < best:
                best, sample = d, s
        if sample is None:
            return
        self._point_menu(lot, readout, col, ph, sample)

    def _point_menu(self, lot, readout, col, phase, sample):
        ph_disp = "" if phase == NO_PHASE else f" [{phase}]"
        m = tk.Menu(self, tearoff=0)
        m.add_command(label=f"{lot}{ph_disp} | Sample {sample} @ {readout}",
                      state="disabled")
        m.add_separator()
        m.add_command(label="Marker 색 변경 (삼각형 표시)",
                      command=lambda: self._change_color(lot, readout, col, phase, sample))
        m.add_command(label="이 Read-out Marker만 삭제",
                      command=lambda: (self.model.delete_point(lot, readout, col, phase, sample),
                                       self._redraw()))
        m.add_command(label="Sample 전체 삭제 (모든 Read-out)",
                      command=lambda: (self.model.delete_sample_all_readouts(lot, col, phase, sample),
                                       self._redraw()))
        m.tk_popup(self.winfo_pointerx(), self.winfo_pointery())

    def _change_color(self, lot, readout, col, phase, sample):
        c = colorchooser.askcolor(parent=self)[1]
        if c:
            self.model.set_color(lot, readout, col, phase, sample, c)
            self._redraw()

    def _set_ylim(self):
        lot, col = self.selected[self.cur_idx]
        cur = self.model.ylim.get((lot, col), (None, None))
        dlg = tk.Toplevel(self)
        dlg.title("Y축 Min/Max")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((lot, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=2, padx=10, pady=(10, 5))
        ttk.Label(dlg, text="Y Max:").grid(row=1, column=0, sticky="e", padx=5)
        vmax = tk.StringVar(value="" if cur[1] is None else str(cur[1]))
        ttk.Entry(dlg, textvariable=vmax, width=15).grid(row=1, column=1, padx=10, pady=3)
        ttk.Label(dlg, text="Y Min:").grid(row=2, column=0, sticky="e", padx=5)
        vmin = tk.StringVar(value="" if cur[0] is None else str(cur[0]))
        ttk.Entry(dlg, textvariable=vmin, width=15).grid(row=2, column=1, padx=10, pady=3)

        def apply():
            try:
                ymin = float(vmin.get())
                ymax = float(vmax.get())
            except ValueError:
                messagebox.showwarning("알림", "숫자를 입력하세요.", parent=dlg)
                return
            if ymax <= ymin:
                messagebox.showwarning("알림", "Y Max는 Y Min보다 커야 합니다.", parent=dlg)
                return
            self.model.set_ylim(lot, col, ymin, ymax)
            dlg.destroy()
            self._redraw()

        def reset():
            self.model.set_ylim(lot, col, None, None)
            dlg.destroy()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(btns, text="적용", command=apply).pack(side="left", padx=5)
        ttk.Button(btns, text="자동(초기화)", command=reset).pack(side="left", padx=5)
        ttk.Button(btns, text="취소", command=dlg.destroy).pack(side="left", padx=5)
        center_window(dlg)

    def _ref_line_dialog(self):
        lot, col = self.selected[self.cur_idx]
        kinds = [("Read-out graph", "line"), ("Delta % graph", "delta"), ("Box plot", "box")]
        # 현재 보고 있는 탭을 기본 종류로
        try:
            cur_tab = self.nb.index(self.nb.select())
        except Exception:
            cur_tab = 0
        kind_var = tk.StringVar(value=kinds[min(cur_tab, 2)][1])

        dlg = tk.Toplevel(self)
        dlg.title("기준선 관리")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((lot, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=4, padx=10, pady=(10, 5))

        ttk.Label(dlg, text="대상 그래프:").grid(row=1, column=0, sticky="e", padx=5)
        kind_cb = ttk.Combobox(dlg, state="readonly", width=16,
                               values=[k[0] for k in kinds])
        kind_cb.current(min(cur_tab, 2))
        kind_cb.grid(row=1, column=1, columnspan=3, sticky="w", pady=2)

        lb = tk.Listbox(dlg, height=5, width=48)
        lb.grid(row=2, column=0, columnspan=4, padx=10, pady=5)

        def cur_kind():
            return kinds[kind_cb.current()][1]

        def refresh():
            lb.delete(0, "end")
            for rl in self.model.ref_lines.get((lot, col, cur_kind()), []):
                lb.insert("end", f"[{rl['axis'].upper()}={rl['value']}] "
                                 f"{rl['name']}  ({rl['color']})")
        kind_cb.bind("<<ComboboxSelected>>", lambda e: refresh())

        ttk.Label(dlg, text="축:").grid(row=3, column=0, sticky="e", padx=5)
        axis_cb = ttk.Combobox(dlg, state="readonly", width=4, values=["Y", "X"])
        axis_cb.current(0)
        axis_cb.grid(row=3, column=1, sticky="w")
        ttk.Label(dlg, text="값:").grid(row=3, column=2, sticky="e", padx=5)
        val_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=val_var, width=10).grid(row=3, column=3, sticky="w")
        ttk.Label(dlg, text="이름:").grid(row=4, column=0, sticky="e", padx=5)
        name_var = tk.StringVar(value="High limit")
        ttk.Entry(dlg, textvariable=name_var, width=14).grid(row=4, column=1, sticky="w")
        color_var = tk.StringVar(value="#d62728")
        color_btn = tk.Button(dlg, text="색 선택", bg=color_var.get(), width=8,
                              command=lambda: _pick_color())
        color_btn.grid(row=4, column=2, columnspan=2, sticky="w", padx=5)

        def _pick_color():
            c = colorchooser.askcolor(color=color_var.get(), parent=dlg)[1]
            if c:
                color_var.set(c)
                color_btn.config(bg=c)

        def add():
            try:
                v = float(val_var.get())
            except ValueError:
                messagebox.showwarning("알림", "값에 숫자를 입력하세요.", parent=dlg)
                return
            k = cur_kind()
            lines = list(self.model.ref_lines.get((lot, col, k), []))
            lines.append(dict(axis=axis_cb.get().lower(), value=v,
                              name=name_var.get().strip(), color=color_var.get()))
            self.model.set_ref_lines(lot, col, k, lines)
            refresh()
            self._redraw()

        def remove():
            sel = lb.curselection()
            if not sel:
                return
            k = cur_kind()
            lines = list(self.model.ref_lines.get((lot, col, k), []))
            del lines[sel[0]]
            self.model.set_ref_lines(lot, col, k, lines)
            refresh()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=5, column=0, columnspan=4, pady=10)
        ttk.Button(btns, text="추가", command=add).pack(side="left", padx=5)
        ttk.Button(btns, text="선택 삭제", command=remove).pack(side="left", padx=5)
        ttk.Button(btns, text="닫기", command=dlg.destroy).pack(side="left", padx=5)
        refresh()
        center_window(dlg)

    def _undo(self):
        if self.model.undo():
            self._redraw()

    def _redo(self):
        if self.model.redo():
            self._redraw()

    # ---- PDF ------------------------------------------------------------------
    def _export_pdf(self):
        if len(self.model.groups) == 1:
            default_name = f"{self.model.reliability}_{self.model.groups[0]}_Data_Analysis.pdf"
        else:
            default_name = f"{self.model.reliability}_Data_Analysis.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF", "*.pdf")],
            confirmoverwrite=False)
        if not path:
            return
        if os.path.exists(path):
            if not messagebox.askyesno(
                    "덮어쓰기 확인",
                    f"동일한 이름의 파일이 이미 있습니다.\n\n{os.path.basename(path)}\n\n덮어쓸까요?",
                    parent=self):
                return
        include_phase = messagebox.askyesno(
            "PDF 구성",
            "Box plot의 Phase별 분석도 PDF에 포함할까요?\n"
            "(기본은 Test item별 분석만 포함됩니다)", parent=self)
        win = tk.Toplevel(self)
        win.title("PDF 생성 중")
        win.geometry("360x90")
        ttk.Label(win, text="PDF Report 생성 중...").pack(pady=8)
        pbar = ttk.Progressbar(win, mode="determinate", length=320)
        pbar.pack(pady=5)
        center_window(win, 360, 90)

        def cb(done, total):
            pbar["maximum"] = total
            pbar["value"] = done
            win.update_idletasks()

        def work():
            try:
                export_pdf(self.model, self.selected, path,
                           include_phase=include_phase, progress_cb=cb)
                self.after(0, lambda: (win.destroy(),
                                       messagebox.showinfo("완료", f"PDF 저장 완료:\n{path}",
                                                           parent=self)))
            except Exception as e:
                err = traceback.format_exc()
                self.after(0, lambda: (win.destroy(),
                                       messagebox.showerror("오류", f"PDF 생성 실패:\n{e}\n\n{err}",
                                                            parent=self)))

        threading.Thread(target=work, daemon=True).start()


# ============================================================================
if __name__ == "__main__":
    app = App()
    app.mainloop()
